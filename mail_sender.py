from openpyxl import load_workbook
from win32com.client import Dispatch
from smtplib import SMTP_SSL
from email import encoders
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart

file = r"list.xlsx"
# select the excel file to pick emails from

wb = load_workbook(file, data_only=True)

sheet = wb.active

outlook = Dispatch('outlook.application')

def send_outlook():
    for i in range (1,sheet.max_row+1):
        mail = outlook.CreateItem(0)
        mail.To = sheet.cell(row=i, column=4).value # Select email column from excel sheet
        mail.CC = ''
        mail.Subject = 'Test Mail'
        mail.Body = 'Message body'
        mail.HTMLBody = '<HTML><BODY><p>Hi.</p><p>This is a test mail.</p><HTML><BODY>'
        attachment  = r"" # path to attachment
        mail.Attachments.Add(attachment)
        mail.Send()


def send_gmail():
    """So you want more features and information eh ? Check out this blog. Pretty good eh?
    https://blog.mailtrap.io/sending-emails-in-python-tutorial-with-code-examples/"""
    for i in range (1,sheet.max_row+1):
        gmail_user = input('Enter your your google Email : ') # enter email
        gmail_password = input('Enter your account password : ') # enter password
        mail_from = gmail_user
        mail_to = sheet.cell(row=i, column=1).value # select the email column in excel sheet
            
        msg = MIMEMultipart()
            
        filename = r"file.pdf" # file path for attachment

        msg['From'] = mail_from
        msg['To'] = mail_to
        msg['Cc'] = ""
        msg['Subject'] = 'Subject' # Subject 

        text_file = r"mail1.txt" # edit the text file to include it in your mail

        with open(text_file) as f:
            text = f.read()

        with open(filename, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
                
        encoders.encode_base64(part)

        part.add_header("Content-Disposition",
                            f"attachment; filename= file.pdf",
                            )
        msg.attach(part)

        body = MIMEText(text)
        msg.attach(body)

        try:
            server = SMTP_SSL('smtp.gmail.com', 465)
            server.login(gmail_user, gmail_password)
        except:
            print("error")

        server.sendmail(mail_from, mail_to, msg.as_string())
        print("mail sent to {}".format(sheet.cell(row=i, column=4).value)) # prints mail ID as a check
        server.close()

send_gmail()
